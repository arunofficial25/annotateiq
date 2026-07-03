import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

# ── Paths ────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT   = os.path.join(BASE_DIR, 'outputs')
os.makedirs(OUTPUT, exist_ok=True)

# ── DB Connection ─────────────────────────────────────────────
from config import SERVER, DATABASE
import urllib
from sqlalchemy import create_engine

params = urllib.parse.quote_plus(
    f"DRIVER={{SQL Server}};"
    f"SERVER={SERVER};"
    f"DATABASE={DATABASE};"
    "Trusted_Connection=yes;"
)
engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
conn   = engine.connect()

print("✅ Connected to database")

# ── Pull Data ─────────────────────────────────────────────────
cost_df = pd.read_sql("""
    SELECT
        p.project_name,
        p.annotation_type,
        p.client_industry,
        p.total_budget_inr,
        COUNT(an.annotation_id)                             AS total_annotations,
        ROUND(p.total_budget_inr / 
              NULLIF(COUNT(an.annotation_id),0), 2)         AS cost_per_annotation,
        SUM(CASE WHEN an.status IN 
            ('rejected','flagged') THEN 1 ELSE 0 END)       AS poor_quality_count,
        ROUND(100.0 * SUM(CASE WHEN an.status IN 
            ('rejected','flagged') THEN 1 ELSE 0 END) 
              / NULLIF(COUNT(an.annotation_id),0), 2)        AS poor_quality_pct,
        ROUND(p.total_budget_inr * (1.0 * SUM(CASE WHEN 
            an.status IN ('rejected','flagged') THEN 1 ELSE 0 END)
              / NULLIF(COUNT(an.annotation_id),0)), 2)       AS wasted_budget_inr
    FROM projects p
    JOIN tasks        t  ON p.project_id  = t.project_id
    JOIN annotations  an ON t.task_id     = an.task_id
    GROUP BY p.project_name, p.annotation_type,
             p.client_industry, p.total_budget_inr
""", conn)

annotator_df = pd.read_sql("""
    SELECT
        a.name, a.region, a.education_level,
        a.employment_type, a.hourly_rate_inr,
        COUNT(an.annotation_id)                              AS total_tasks,
        ROUND(SUM(an.time_taken_minutes)/60.0, 2)           AS total_hours,
        ROUND((SUM(an.time_taken_minutes)/60.0) 
               * a.hourly_rate_inr, 2)                      AS total_earned_inr,
        ROUND(100.0 * SUM(CASE WHEN an.status = 'approved' 
              THEN 1 ELSE 0 END) 
              / NULLIF(COUNT(an.annotation_id),0), 2)        AS approval_rate_pct
    FROM annotators a
    JOIN annotations an ON a.annotator_id = an.annotator_id
    GROUP BY a.name, a.region, a.education_level,
             a.employment_type, a.hourly_rate_inr
""", conn)

workforce_df = pd.read_sql("""
    SELECT
        wi.economic_region, wi.displacement_risk,
        wi.avg_worker_literacy, wi.retraining_feasibility,
        SUM(wi.jobs_currently_needed)   AS current_jobs,
        SUM(wi.projected_jobs_5yr)      AS projected_jobs,
        SUM(wi.jobs_currently_needed) 
        - SUM(wi.projected_jobs_5yr)    AS jobs_at_risk
    FROM workforce_impact wi
    GROUP BY wi.economic_region, wi.displacement_risk,
             wi.avg_worker_literacy, wi.retraining_feasibility
""", conn)

print("✅ Data loaded")

# ── Styles ────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
RED        = "C00000"
ORANGE     = "ED7D31"
GREEN      = "70AD47"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"

def header_style(cell, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def border_all(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_table(ws, df, start_row, start_col, header_bg=MID_BLUE):
    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, 
                       column=start_col + col_idx, 
                       value=col_name.replace('_',' ').title())
        header_style(cell, bg=header_bg)
        border_all(cell)
    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row):
            cell = ws.cell(row=start_row + row_idx + 1,
                           column=start_col + col_idx,
                           value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", 
                        fgColor=LIGHT_GREY if row_idx % 2 == 0 else WHITE)
            border_all(cell)
    # Auto width
    for col_idx in range(len(df.columns)):
        col_letter = get_column_letter(start_col + col_idx)
        ws.column_dimensions[col_letter].width = 22

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — Cover
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Cover"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("B2:H2")
title_cell = ws1["B2"]
title_cell.value    = "AnnotateIQ — Cost & Workforce Impact Report"
title_cell.font     = Font(size=20, bold=True, color=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center")

ws1.merge_cells("B3:H3")
sub = ws1["B3"]
sub.value     = "AI Annotation Analytics | India Context | 2023–2024"
sub.font      = Font(size=12, italic=True, color=MID_BLUE)
sub.alignment = Alignment(horizontal="center")

ws1.merge_cells("B5:H5")
tag = ws1["B5"]
tag.value     = '"Behind every AI model is a human whose job it trained away."'
tag.font      = Font(size=11, italic=True, color=RED)
tag.alignment = Alignment(horizontal="center")

labels = [
    ("Sheet", "Content"),
    ("Project Cost Analysis", "Budget, cost per annotation, wasted spend"),
    ("Annotator Economics",   "Earnings, hours, approval rates"),
    ("Workforce Impact",      "Job displacement by region & literacy"),
    ("Summary KPIs",          "Key metrics at a glance"),
]
for i, (sheet, content) in enumerate(labels):
    row = 8 + i
    c1 = ws1.cell(row=row, column=2, value=sheet)
    c2 = ws1.cell(row=row, column=5, value=content)
    if i == 0:
        header_style(c1, bg=DARK_BLUE)
        header_style(c2, bg=DARK_BLUE)
    else:
        c1.font = Font(bold=True, color=MID_BLUE)
        c2.font = Font(color="444444")
    border_all(c1)
    border_all(c2)

ws1.row_dimensions[2].height = 35
ws1.row_dimensions[3].height = 20
ws1.row_dimensions[5].height = 20

# ══════════════════════════════════════════════════════════════
# SHEET 2 — Project Cost Analysis
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Project Cost Analysis")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:I1")
t = ws2["A1"]
t.value     = "Project Cost Analysis"
t.font      = Font(size=14, bold=True, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

write_table(ws2, cost_df, start_row=3, start_col=1)

# Conditional color on wasted_budget column (last col)
waste_col = len(cost_df.columns)
for row in range(4, 4 + len(cost_df)):
    cell = ws2.cell(row=row, column=waste_col)
    val  = cell.value or 0
    if val > 200000:
        cell.fill = PatternFill("solid", fgColor="FFCCCC")
        cell.font = Font(color=RED, bold=True)
    elif val > 100000:
        cell.fill = PatternFill("solid", fgColor="FFE5CC")
        cell.font = Font(color=ORANGE, bold=True)

# ══════════════════════════════════════════════════════════════
# SHEET 3 — Annotator Economics
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Annotator Economics")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:I1")
t = ws3["A1"]
t.value     = "Annotator Economics — Earnings & Performance"
t.font      = Font(size=14, bold=True, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

write_table(ws3, annotator_df, start_row=3, start_col=1)

# Color approval rate column
apr_col = len(annotator_df.columns)
for row in range(4, 4 + len(annotator_df)):
    cell = ws3.cell(row=row, column=apr_col)
    val  = cell.value or 0
    if val == 100:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="276221", bold=True)
    elif val == 0:
        cell.fill = PatternFill("solid", fgColor="FFCCCC")
        cell.font = Font(color=RED, bold=True)

# ══════════════════════════════════════════════════════════════
# SHEET 4 — Workforce Impact
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Workforce Impact")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:G1")
t = ws4["A1"]
t.value     = "Workforce Displacement by Region & Literacy"
t.font      = Font(size=14, bold=True, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

write_table(ws4, workforce_df, start_row=3, start_col=1)

# Color displacement risk column (col 2)
for row in range(4, 4 + len(workforce_df)):
    cell = ws4.cell(row=row, column=2)
    val  = str(cell.value).lower()
    if val == 'critical':
        cell.fill = PatternFill("solid", fgColor="FFCCCC")
        cell.font = Font(color=RED, bold=True)
    elif val == 'high':
        cell.fill = PatternFill("solid", fgColor="FFE5CC")
        cell.font = Font(color=ORANGE, bold=True)
    elif val == 'medium':
        cell.fill = PatternFill("solid", fgColor="FFFF99")
        cell.font = Font(color="7F6000", bold=True)
    elif val == 'low':
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="276221", bold=True)

# ══════════════════════════════════════════════════════════════
# SHEET 5 — Summary KPIs
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary KPIs")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("B1:F1")
t = ws5["B1"]
t.value     = "AnnotateIQ — Executive Summary KPIs"
t.font      = Font(size=14, bold=True, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

total_budget     = cost_df['total_budget_inr'].sum()
total_wasted     = cost_df['wasted_budget_inr'].sum()
total_annotations= cost_df['total_annotations'].sum()
avg_cost         = cost_df['cost_per_annotation'].mean()
total_jobs_now   = workforce_df['current_jobs'].sum()
total_jobs_5yr   = workforce_df['projected_jobs'].sum()
jobs_lost        = workforce_df['jobs_at_risk'].sum()

kpis = [
    ("Total Project Budget (INR)",      f"₹{total_budget:,.0f}",    MID_BLUE),
    ("Estimated Wasted Budget (INR)",   f"₹{total_wasted:,.0f}",    RED),
    ("Total Annotations Completed",     f"{total_annotations:,}",    MID_BLUE),
    ("Avg Cost per Annotation (INR)",   f"₹{avg_cost:,.2f}",         MID_BLUE),
    ("Current Annotation Jobs",         f"{total_jobs_now:,}",       MID_BLUE),
    ("Projected Jobs in 5 Years",       f"{total_jobs_5yr:,}",       ORANGE),
    ("Jobs at Risk of Displacement",    f"{jobs_lost:,}",            RED),
    ("Displacement Rate",               f"{100*jobs_lost/total_jobs_now:.1f}%", RED),
]

for i, (label, value, color) in enumerate(kpis):
    row = 3 + i * 2
    ws5.merge_cells(f"B{row}:D{row}")
    ws5.merge_cells(f"E{row}:F{row}")
    lc = ws5[f"B{row}"]
    vc = ws5[f"E{row}"]
    lc.value     = label
    vc.value     = value
    lc.font      = Font(bold=True, size=11, color="333333")
    vc.font      = Font(bold=True, size=13, color=color)
    vc.alignment = Alignment(horizontal="center")
    lc.fill      = PatternFill("solid", fgColor=LIGHT_GREY)
    vc.fill      = PatternFill("solid", fgColor="EBF3FF")
    border_all(lc)
    border_all(vc)
    ws5.row_dimensions[row].height = 22

# ── Save ──────────────────────────────────────────────────────
path = os.path.join(OUTPUT, 'AnnotateIQ_Cost_Report.xlsx')
wb.save(path)
conn.close()
print(f"✅ Excel workbook saved → {path}")